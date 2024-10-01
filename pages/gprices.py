import openpyxl
import os
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivymd.uix.button import MDRaisedButton
from kivymd.uix.label import MDLabel
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from styles import Styles

class GPricesPage(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.create_ui_elements()
        self.input_file = None  # Изначально файл не выбран

    def process_file(self, instance):
        try:
            if not hasattr(self, 'input_file') or not os.path.isfile(self.input_file):
                self.label.text = "Файл не найден. Пожалуйста, выберите файл."
                return
            
            # Определяем путь к выходному файлу
            output_file = os.path.join(os.path.dirname(self.input_file), 'output.xlsx')

            # Проверяем, существует ли выходной файл и выводим сообщение
            file_exists = os.path.isfile(output_file)

            # Читаем данные из файла Excel
            wb_input = openpyxl.load_workbook(self.input_file)
            ws_input = wb_input.active

            # Создаем новый Workbook для выходного файла
            wb_output = openpyxl.Workbook()
            ws_output = wb_output.active

            # Заголовки для выходного файла
            headers = ['DEST', 'COUNTRY CODE', 'ROLLUP', 'RATE', 'MC', 'CI', 'FT']
            ws_output.append(headers)

            # Обрабатываем каждую строку в исходном файле
            for row in ws_input.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
                dest = row[0]
                country_code = str(row[1])  # Оставляем как строку
                rollup = str(row[2]) if row[2] is not None else ''  # Проверка на None
                rate = f"{float(row[3]):.4f}"  # Форматирование значения RATE с 4 знаками после запятой

                # Создаем список для хранения обработанных данных
                processed_data = []

                # Проверяем, есть ли значения в ROLLUP
                if rollup:
                    # Разделяем значения по запятой и тире
                    rollup_values = []
                    for part in rollup.split(','):
                        part = part.strip()
                        if '-' in part:  # Обработка диапазонов
                            start_end = part.split('-')
                            if len(start_end) == 2 and all(x.strip().isdigit() for x in start_end):
                                start, end = map(int, start_end)
                                rollup_values.extend(range(start, end + 1))
                        else:  # Обработка отдельных значений
                            if part.isdigit():  # Проверка на целое число
                                rollup_values.append(int(part))

                    # Формируем строки с отдельными COUNTRY CODE и ROLLUP
                    for value in rollup_values:
                        processed_data.append([dest, country_code, value, rate, 1, 1, 0])  # Добавляем значение ROLLUP

                else:
                    # Если ROLLUP пустой или None, добавляем строку без изменений (исключая комментарии и дату)
                    processed_data.append([dest, country_code, '', rate, 1, 1, 0])  # Пустое значение для ROLLUP

                # Записываем обработанные данные в выходной файл с проверкой на ведущие нули
                for data_row in processed_data:
                    dest_val, country_code_val, rollup_val, rate_val, mc_val, ci_val, ft_val = data_row
                    
                    # Преобразуем значение ROLLUP в строку и проверяем наличие ведущих нулей
                    if isinstance(rollup_val, int):
                        rollup_val_str = str(rollup_val)
                        original_rollups = [part.strip() for part in rollup.split(',')]
                        if any(r.lstrip('0') == rollup_val_str for r in original_rollups):
                            rollup_val_str = str(rollup_val)  # Сохраняем без добавления нуля
                        else:
                            rollup_val_str = str(rollup_val).zfill(2)  # Добавляем ведущий ноль если нужно
                    else:
                        rollup_val_str = ''

                    ws_output.append([dest_val, country_code_val, rollup_val_str, rate_val, mc_val, ci_val, ft_val])

            # Сохраняем выходной файл
            wb_output.save(output_file)

            self.label.text += f"\nГотово. Файл сохранен как: {output_file}"
            if file_exists:
                self.label.text += f"\nФайл {output_file} был перезаписан."
        except Exception as e:
            self.label.text = f"Ошибка: {str(e)}, файл некорректен"
        
    def create_ui_elements(self):
        main_layout = BoxLayout(
            orientation="vertical",
            padding="20dp",
            spacing="20dp"
        )
        self.add_widget(main_layout)

        # Заголовок
        self.label = MDLabel(
            text="Скрипт для обработки говнопрайса Билайн\nВыберите файл input.xlsx для обработки\nДля корректной работы скрипта вам нужно ТОЛЬКО переименовать файл в input.xlsx",
            halign="center",
            size_hint_y=None,
            height="200dp"
        )
        main_layout.add_widget(self.label)

        # Поле ввода для отображения выбранного файла
        self.file_display = MDLabel(
            text="После выбора файла и жмяк по кнопке обработать, дождитесь надписи `Готово...`",
            halign="center",
            size_hint_y=None,
            height="200dp"
        )
        main_layout.add_widget(self.file_display)

        button_layout = BoxLayout(
            orientation="horizontal",
            spacing="20dp",
            size_hint_y=None,
            height="60dp"
        )
        main_layout.add_widget(button_layout)

        # Кнопка выбора файла
        self.choose_file_button = MDRaisedButton(
            text="Выбрать файл",
            on_press=self.choose_file,
            font_name=Styles.ENTRY_FONT,
            theme_text_color="Custom",
            text_color=Styles.BUTTON_TEXT_COLOR,
            md_bg_color=Styles.BUTTON_COLOR,
        )
        button_layout.add_widget(self.choose_file_button)

        # Кнопка обработать
        self.process_button = MDRaisedButton(
            text="Обработать",
            on_press=self.process_file,
            font_name=Styles.ENTRY_FONT,
            theme_text_color="Custom",
            text_color=Styles.BUTTON_TEXT_COLOR,
            md_bg_color=Styles.BUTTON_COLOR,
        )
        button_layout.add_widget(self.process_button)

        # Кнопка назад
        self.back_button = MDRaisedButton(
            text="Назад",
            on_press=self.back_button_pressed,
            font_name=Styles.ENTRY_FONT,
            theme_text_color="Custom",
            text_color=Styles.BUTTON_TEXT_COLOR,
            md_bg_color=Styles.BUTTON_COLOR,
        )
        button_layout.add_widget(self.back_button)
        
#Подвязка скрипта к текстовым полям и кнопкам
    def back_button_pressed(self, instance):
        self.manager.current = "main"
        
    def choose_file(self, instance):
        # Открываем системный проводник для выбора файла
        Tk().withdraw()  # Скрываем главное окно Tkinter
        self.input_file = askopenfilename(title="Выберите файл", filetypes=[("Excel files", "*.xlsx")])
        
        if self.input_file:
            self.label.text = f"Выбранный файл: {self.input_file}"
        else:
            self.label.text = "После выбора файла и жмяк по кнопке обработать, дождитесь надписи `Готово...`"